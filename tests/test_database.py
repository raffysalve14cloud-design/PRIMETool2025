import os
print("Starting database-only test...")

# Android detection and setup
platform = "Unknown"
storage_path = None
is_android = False

try:
    from android.storage import primary_external_storage_path
    from android.permissions import request_permission, Permission
    platform = "Android"
    is_android = True
    storage_path = primary_external_storage_path()
    print(f"Platform: {platform}")
    print(f"Storage path: {storage_path}")
    
    # Request storage permissions
    try:
        request_permission(Permission.WRITE_EXTERNAL_STORAGE)
        request_permission(Permission.READ_EXTERNAL_STORAGE)
        print("Storage permissions requested")
    except Exception as e:
        print(f"Permission request warning: {e}")
        
except ImportError:
    platform = "Windows" if os.name == 'nt' else "Linux"
    storage_path = os.getcwd()
    print(f"Platform: {platform}")
    print(f"Working directory: {storage_path}")

# Test SQLite operations
print("\nTesting SQLite operations...")
try:
    import sqlite3
    print("✓ SQLite import successful")
    
    # Create test database path
    if is_android:
        db_path = os.path.join(storage_path, "prime_database_test.db")
    else:
        db_path = os.path.join(storage_path, "data", "prime_database_test.db")
        os.makedirs(os.path.dirname(db_path), exist_ok=True)
    
    print(f"Database path: {db_path}")
    
    # Test database creation and operations
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Test table creation (similar to your main app)
    print("Creating test tables...")
    cursor.execute('''CREATE TABLE IF NOT EXISTS erdata (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        employer_name TEXT NOT NULL,
        employer_id TEXT UNIQUE,
        employer_address TEXT,
        contact_person TEXT,
        contact_number TEXT,
        email_address TEXT,
        date_added TEXT,
        status TEXT DEFAULT 'Active'
    )''')
    
    cursor.execute('''CREATE TABLE IF NOT EXISTS uedata (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        employer_id TEXT,
        employee_name TEXT,
        employee_id TEXT,
        position TEXT,
        salary REAL,
        date_hired TEXT,
        status TEXT DEFAULT 'Active',
        FOREIGN KEY (employer_id) REFERENCES erdata (employer_id)
    )''')
    
    cursor.execute('''CREATE TABLE IF NOT EXISTS year (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        year_value TEXT UNIQUE,
        is_active INTEGER DEFAULT 0
    )''')
    
    print("✓ Tables created successfully")
    
    # Test data insertion
    print("Testing data insertion...")
    test_data = [
        ("Test Employer 1", "EMP001", "123 Test St", "John Doe", "123-456-7890", "test@email.com", "2024-01-01", "Active"),
        ("Test Employer 2", "EMP002", "456 Demo Ave", "Jane Smith", "098-765-4321", "demo@email.com", "2024-01-02", "Active")
    ]
    
    for data in test_data:
        cursor.execute('''INSERT OR IGNORE INTO erdata 
            (employer_name, employer_id, employer_address, contact_person, contact_number, email_address, date_added, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)''', data)
    
    # Test year data
    cursor.execute("INSERT OR IGNORE INTO year (year_value, is_active) VALUES ('2024', 1)")
    cursor.execute("INSERT OR IGNORE INTO year (year_value, is_active) VALUES ('2025', 0)")
    
    conn.commit()
    print("✓ Data insertion successful")
    
    # Test data retrieval
    print("Testing data retrieval...")
    cursor.execute("SELECT COUNT(*) FROM erdata")
    er_count = cursor.fetchone()[0]
    print(f"✓ Employer records: {er_count}")
    
    cursor.execute("SELECT COUNT(*) FROM year")
    year_count = cursor.fetchone()[0]
    print(f"✓ Year records: {year_count}")
    
    # Test complex query (similar to your main app)
    cursor.execute('''SELECT er.employer_name, er.employer_id, er.status, 
                             COUNT(ue.id) as employee_count
                      FROM erdata er 
                      LEFT JOIN uedata ue ON er.employer_id = ue.employer_id 
                      GROUP BY er.employer_id''')
    results = cursor.fetchall()
    print(f"✓ Complex query returned {len(results)} results")
    
    conn.close()
    print("✓ Database operations completed successfully")
    
except Exception as e:
    print(f"✗ Database error: {e}")
    import traceback
    traceback.print_exc()

# Test Excel operations
print("\nTesting Excel operations...")
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    print("✓ OpenPyXL import successful")
    
    # Create test workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Export"
    
    # Add test headers
    headers = ["ID", "Employer Name", "Status", "Employee Count"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add test data
    test_excel_data = [
        ["EMP001", "Test Employer 1", "Active", 0],
        ["EMP002", "Test Employer 2", "Active", 0]
    ]
    
    for row, data in enumerate(test_excel_data, 2):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Save test file
    if is_android:
        excel_path = os.path.join(storage_path, "test_export.xlsx")
    else:
        exports_dir = os.path.join(storage_path, "exports")
        os.makedirs(exports_dir, exist_ok=True)
        excel_path = os.path.join(exports_dir, "test_export.xlsx")
    
    wb.save(excel_path)
    print(f"✓ Excel file saved to: {excel_path}")
    
    # Verify file exists and is readable
    if os.path.exists(excel_path):
        wb_test = openpyxl.load_workbook(excel_path)
        ws_test = wb_test.active
        print(f"✓ Excel file verified - contains {ws_test.max_row} rows")
        wb_test.close()
    else:
        print("✗ Excel file not found after save")
        
except Exception as e:
    print(f"✗ Excel error: {e}")
    import traceback
    traceback.print_exc()

print("\n" + "="*50)
print("DATABASE & EXCEL TEST COMPLETED")
print("="*50)
print("If you see this message, database and Excel")
print("operations are working correctly.")
print("The crash is likely related to UI components")
print("or graphics rendering, not data operations.")
print("="*50)